"""
Facial Expression Recognition - Image & Video processor
Detects faces (OpenCV Haar cascade) and classifies expression (YOLOv8-cls trained on FER2013).
Labels are displayed in Indonesian (see LABELS_ID below); change to taste.

FEATURES:
- For VIDEOS: a bottom-left overlay is burned into every output frame showing
  elapsed video time and accumulated time per detected expression (only
  expressions seen so far are listed). Creates report/<video_name>/ containing
  summary.xlsx (totals + pie/bar charts) and timeline.xlsx (per-frame log).
- For IMAGES: no time overlay is drawn (a single still frame has no duration
  to accumulate). Creates report/<image_name>/detections.xlsx listing every
  detected face's expression and confidence.

ACCUMULATION RULE for videos (read this before trusting the numbers):
  Time is accumulated PER FACE, not per frame. If 3 people are smiling in
  the same frame, that frame's duration is added to "happy" three times
  (once per face) -- this answers "how much total face-time was spent on
  each expression across everyone in the video". Set
  USE_PER_FACE_ACCUMULATION to False below if you instead want each frame
  to count at most once per expression, no matter how many faces share it.
  Video time is driven by the video's own FPS metadata, not wall-clock
  processing speed, so the report reflects in-video time correctly even
  though processing itself may run slower or faster than real-time.

Usage:
    python detect_image_video.py --source path/to/image.jpg
    python detect_image_video.py --source path/to/video.mp4
    python detect_image_video.py --source path/to/folder_of_images
    python detect_image_video.py --source clip.mp4 --report-dir my_reports
"""

import argparse
import os
from collections import defaultdict
from datetime import datetime

import cv2
from ultralytics import YOLO

# ---- CONFIG ----------------------------------------------------------

MODEL_PATH = "fer2013_best.pt"  # path to your trained YOLO classification model

USE_PER_FACE_ACCUMULATION = True

# Maps the model's English class names to Indonesian for display only.
LABELS_ID = {
    "angry":    "MARAH",
    "disgust":  "JIJIK",
    "fear":     "TAKUT",
    "happy":    "SENANG",
    "neutral":  "NETRAL",
    "sad":      "SEDIH",
    "surprise": "TERKEJUT",
}

EMOTION_COLORS = {
    "angry":    (0, 0, 255),     # red
    "disgust":  (0, 140, 255),   # orange
    "fear":     (128, 0, 128),   # purple
    "happy":    (0, 255, 0),     # green
    "neutral":  (200, 200, 200), # light gray
    "sad":      (255, 0, 0),     # blue
    "surprise": (0, 255, 255),   # yellow
}

BOX_THICKNESS = 4
FONT = cv2.FONT_HERSHEY_SIMPLEX
FONT_SCALE = 1.0
FONT_THICKNESS = 2
PADDING = 10  # extra pixels added around the detected face crop before classification

OVERLAY_FONT_SCALE = 0.8
OVERLAY_FONT_THICKNESS = 2
OVERLAY_LINE_HEIGHT = 32
OVERLAY_MARGIN = 16

# ------------------------------------------------------------------------

face_cascade = cv2.CascadeClassifier(
    cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
)

print("Loading expression model...")
model = YOLO(MODEL_PATH)


def classify_face(face_img):
    """Run the YOLO classifier on a cropped face image. Returns (label, confidence)."""
    if face_img.size == 0:
        return None, 0.0
    results = model.predict(face_img, verbose=False)
    probs = results[0].probs
    top1_idx = int(probs.top1)
    conf = float(probs.top1conf)
    label = model.names[top1_idx]
    return label, conf


def draw_label_box(frame, x, y, w, h, label, conf):
    color = EMOTION_COLORS.get(label, (255, 255, 255))
    display_text = LABELS_ID.get(label, label.upper())
    text = f"{display_text} {conf*100:.0f}%"

    cv2.rectangle(frame, (x, y), (x + w, y + h), color, BOX_THICKNESS)

    (text_w, text_h), baseline = cv2.getTextSize(text, FONT, FONT_SCALE, FONT_THICKNESS)
    label_y = max(0, y - text_h - 15)
    cv2.rectangle(frame, (x, label_y), (x + text_w + 12, label_y + text_h + 15), color, -1)

    brightness = sum(color) / 3
    text_color = (0, 0, 0) if brightness > 150 else (255, 255, 255)
    cv2.putText(
        frame, text, (x + 6, label_y + text_h + 5),
        FONT, FONT_SCALE, text_color, FONT_THICKNESS, cv2.LINE_AA
    )


def format_duration(seconds):
    seconds = int(seconds)
    h, rem = divmod(seconds, 3600)
    m, s = divmod(rem, 60)
    if h > 0:
        return f"{h:02d}:{m:02d}:{s:02d}"
    return f"{m:02d}:{s:02d}"


def draw_overlay(frame, elapsed_seconds, expression_totals):
    """Bottom-left overlay: accumulated time per detected expression + elapsed time."""
    h_img, w_img = frame.shape[:2]

    detected = [(label, secs) for label, secs in expression_totals.items() if secs > 0]
    detected.sort(key=lambda item: item[1], reverse=True)

    lines = []
    for label, secs in detected:
        display_text = LABELS_ID.get(label, label.upper())
        lines.append((display_text, EMOTION_COLORS.get(label, (255, 255, 255)), format_duration(secs)))
    lines.append(("Time elapsed", (255, 255, 255), format_duration(elapsed_seconds)))

    box_w = 320
    box_h = OVERLAY_MARGIN * 2 + OVERLAY_LINE_HEIGHT * len(lines)
    box_x0 = OVERLAY_MARGIN
    box_y1 = h_img - OVERLAY_MARGIN
    box_y0 = box_y1 - box_h
    box_x1 = box_x0 + box_w

    overlay = frame.copy()
    cv2.rectangle(overlay, (box_x0, box_y0), (box_x1, box_y1), (0, 0, 0), -1)
    cv2.addWeighted(overlay, 0.55, frame, 0.45, 0, dst=frame)

    text_y = box_y0 + OVERLAY_MARGIN + 20
    for label_text, color, value_text in lines:
        line = f"{label_text}: {value_text}"
        cv2.putText(frame, line, (box_x0 + 14, text_y),
                    FONT, OVERLAY_FONT_SCALE, color, OVERLAY_FONT_THICKNESS, cv2.LINE_AA)
        text_y += OVERLAY_LINE_HEIGHT


def save_video_summary(path, source_name, expression_totals, timeline):
    """Standalone xlsx: Summary totals + pie/bar charts for a processed video."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.chart import PieChart, BarChart, Reference

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Facial Expression Video Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Source video"
    ws["B2"] = source_name
    ws["A3"] = "Generated"
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A4"] = "Total video duration (seconds)"
    ws["B4"] = round(timeline[-1]["elapsed"], 1) if timeline else 0

    header_row = 6
    ws.cell(row=header_row, column=1, value="Expression").font = Font(bold=True)
    ws.cell(row=header_row, column=2, value="Seconds").font = Font(bold=True)
    ws.cell(row=header_row, column=3, value="% of total face-time").font = Font(bold=True)
    for c in (1, 2, 3):
        ws.cell(row=header_row, column=c).fill = PatternFill("solid", start_color="DDDDDD")

    sorted_items = sorted(
        ((LABELS_ID.get(k, k.upper()), v) for k, v in expression_totals.items() if v > 0),
        key=lambda x: x[1], reverse=True
    )

    first_data_row = header_row + 1
    row = first_data_row
    for label, secs in sorted_items:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=round(secs, 1))
        row += 1
    last_data_row = row - 1

    if last_data_row >= first_data_row:
        total_cell = f"SUM(B{first_data_row}:B{last_data_row})"
        for r in range(first_data_row, last_data_row + 1):
            ws.cell(row=r, column=3, value=f"=B{r}/{total_cell}")
            ws.cell(row=r, column=3).number_format = "0.0%"

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 18

        pie = PieChart()
        pie.title = "Share of detected expression time"
        data = Reference(ws, min_col=2, min_row=header_row, max_row=last_data_row)
        cats = Reference(ws, min_col=1, min_row=first_data_row, max_row=last_data_row)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(cats)
        ws.add_chart(pie, "E6")

        bar = BarChart()
        bar.title = "Accumulated seconds per expression"
        bar.y_axis.title = "Seconds"
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        ws.add_chart(bar, "E22")
    else:
        ws.cell(row=first_data_row, column=1, value="No expressions detected in this video.")

    wb.save(path)
    print(f"Summary report saved: {path}")


def save_video_timeline(path, timeline):
    """Standalone xlsx: per-frame timeline for a processed video."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws2 = wb.active
    ws2.title = "Timeline"
    expr_keys = list(LABELS_ID.keys())
    headers = ["Elapsed (s)"] + [LABELS_ID[k] for k in expr_keys] + ["Faces detected (this frame)"]
    ws2.append(headers)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", start_color="DDDDDD")

    for sample in timeline:
        row = [round(sample["elapsed"], 2)]
        for k in expr_keys:
            row.append(round(sample["deltas"].get(k, 0.0), 3))
        row.append(sample["face_count"])
        ws2.append(row)

    for i in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 14

    wb.save(path)
    print(f"Timeline report saved: {path}")


def save_image_detections(path, source_name, face_results):
    """Standalone xlsx: one row per detected face in a processed image."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Detections"
    ws["A1"] = "Facial Expression Image Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Source image"
    ws["B2"] = source_name
    ws["A3"] = "Generated"
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A4"] = "Faces detected"
    ws["B4"] = len(face_results)

    header_row = 6
    headers = ["Face #", "Expression", "Confidence %"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", start_color="DDDDDD")

    for i, (label, conf) in enumerate(face_results, start=1):
        r = header_row + i
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=LABELS_ID.get(label, label.upper()))
        ws.cell(row=r, column=3, value=round(conf * 100, 1))

    for col, width in zip("ABC", (10, 16, 14)):
        ws.column_dimensions[col].width = width

    wb.save(path)
    print(f"Report saved: {path}")


def process_image(path, output_dir, report_dir):
    img = cv2.imread(path)
    if img is None:
        print(f"Could not read image: {path}")
        return

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(
        gray, scaleFactor=1.1, minNeighbors=6, minSize=(60, 60)
    )

    h_img, w_img = img.shape[:2]
    face_results = []

    for (x, y, w, h) in faces:
        x0 = max(0, x - PADDING)
        y0 = max(0, y - PADDING)
        x1 = min(w_img, x + w + PADDING)
        y1 = min(h_img, y + h + PADDING)

        face_crop = img[y0:y1, x0:x1]
        label, conf = classify_face(face_crop)

        if label is not None:
            draw_label_box(img, x0, y0, x1 - x0, y1 - y0, label, conf)
            face_results.append((label, conf))

    os.makedirs(output_dir, exist_ok=True)
    base = os.path.splitext(os.path.basename(path))[0]
    out_path = os.path.join(output_dir, "out_" + os.path.basename(path))
    cv2.imwrite(out_path, img)
    print(f"Saved: {out_path}")

    session_dir = os.path.join(report_dir, base)
    os.makedirs(session_dir, exist_ok=True)
    save_image_detections(os.path.join(session_dir, "detections.xlsx"), os.path.basename(path), face_results)


def process_video(path, output_dir, report_dir):
    cap = cv2.VideoCapture(path)
    if not cap.isOpened():
        print(f"Could not open video: {path}")
        return

    fps = cap.get(cv2.CAP_PROP_FPS) or 25
    dt = 1.0 / fps
    w = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    h = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))

    os.makedirs(output_dir, exist_ok=True)
    base = os.path.splitext(os.path.basename(path))[0]
    out_path = os.path.join(output_dir, "out_" + base + ".mp4")
    fourcc = cv2.VideoWriter_fourcc(*"mp4v")
    writer = cv2.VideoWriter(out_path, fourcc, fps, (w, h))

    expression_totals = defaultdict(float)
    timeline = []

    frame_count = 0
    print("Processing video... this may take a while.")
    while True:
        ret, frame = cap.read()
        if not ret:
            break

        elapsed = frame_count * dt

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(
            gray, scaleFactor=1.1, minNeighbors=6, minSize=(60, 60)
        )

        h_img, w_img = frame.shape[:2]
        sample_deltas = defaultdict(float)

        for (x, y, w_box, h_box) in faces:
            x0 = max(0, x - PADDING)
            y0 = max(0, y - PADDING)
            x1 = min(w_img, x + w_box + PADDING)
            y1 = min(h_img, y + h_box + PADDING)

            face_crop = frame[y0:y1, x0:x1]
            label, conf = classify_face(face_crop)
            if label is None:
                continue
            draw_label_box(frame, x0, y0, x1 - x0, y1 - y0, label, conf)

            if USE_PER_FACE_ACCUMULATION:
                expression_totals[label] += dt
                sample_deltas[label] += dt
            else:
                sample_deltas[label] = dt

        if not USE_PER_FACE_ACCUMULATION:
            for label in sample_deltas:
                expression_totals[label] += dt

        timeline.append({
            "elapsed": elapsed,
            "deltas": dict(sample_deltas),
            "face_count": len(faces),
        })

        draw_overlay(frame, elapsed, expression_totals)
        writer.write(frame)

        frame_count += 1
        if frame_count % 30 == 0:
            print(f"  processed {frame_count} frames...")

    cap.release()
    writer.release()
    print(f"Saved: {out_path}")

    session_dir = os.path.join(report_dir, base)
    os.makedirs(session_dir, exist_ok=True)
    save_video_summary(os.path.join(session_dir, "summary.xlsx"), os.path.basename(path), expression_totals, timeline)
    save_video_timeline(os.path.join(session_dir, "timeline.xlsx"), timeline)


def main():
    parser = argparse.ArgumentParser(description="Facial expression recognition for images/videos")
    parser.add_argument("--source", required=True, help="Path to an image, a video, or a folder of images")
    parser.add_argument("--output", default="output", help="Output directory for processed images/videos")
    parser.add_argument(
        "--report-dir", default="report",
        help="Base folder for reports. Each processed file gets its own subfolder here "
             "(report/<filename>/) containing its .xlsx report(s). Default: ./report"
    )
    args = parser.parse_args()

    src = args.source
    image_exts = (".jpg", ".jpeg", ".png", ".bmp", ".webp")
    video_exts = (".mp4", ".avi", ".mov", ".mkv")

    if os.path.isdir(src):
        for fname in os.listdir(src):
            fpath = os.path.join(src, fname)
            if fname.lower().endswith(image_exts):
                process_image(fpath, args.output, args.report_dir)
            elif fname.lower().endswith(video_exts):
                process_video(fpath, args.output, args.report_dir)
    elif src.lower().endswith(image_exts):
        process_image(src, args.output, args.report_dir)
    elif src.lower().endswith(video_exts):
        process_video(src, args.output, args.report_dir)
    else:
        print("Unsupported file type. Use an image, video, or folder.")


if __name__ == "__main__":
    main()