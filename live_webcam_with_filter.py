"""
Usage:
    python live_webcam.py
    python live_webcam.py --camera 1
    python live_webcam.py --fullscreen
    python live_webcam.py --width 1280 --height 720
    python live_webcam.py --list-cameras
    python live_webcam.py --report-dir my_reports   (default: ./report)

Press 'q' to quit (saves report), 'f' to toggle fullscreen, 'e' to save report without quitting.
"""

import argparse
import os
import sys
import time
from collections import defaultdict
from datetime import datetime

import cv2
from ultralytics import YOLO

MODEL_PATH = "fer2013_best.pt"

USE_PER_FACE_ACCUMULATION = True

# Only these expressions are detected/shown/counted. The model itself still
# predicts all classes it was trained on; anything not in this set is simply
# ignored (no box, no label, no time accumulated).
ALLOWED_EXPRESSIONS = {"neutral", "sad", "surprise"}

LABELS_ID = {
    "neutral":  "SERIUS",
    "sad":      "NGANTUK",
    "surprise": "NGOBROL",
}

EMOTION_COLORS = {
    "neutral":  (200, 200, 200),
    "sad":      (255, 0, 0),
    "surprise": (0, 255, 255),
}

BOX_THICKNESS = 4
FONT = cv2.FONT_HERSHEY_SIMPLEX
FONT_SCALE = 1.0
FONT_THICKNESS = 2
PADDING = 10

OVERLAY_FONT_SCALE = 0.8
OVERLAY_FONT_THICKNESS = 2
OVERLAY_LINE_HEIGHT = 32
OVERLAY_MARGIN = 16

face_cascade = cv2.CascadeClassifier(
    cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
)

print("Loading expression model...")
model = YOLO(MODEL_PATH)


def classify_face(face_img):
    if face_img.size == 0:
        return None, 0.0
    results = model.predict(face_img, verbose=False)
    probs = results[0].probs
    label = model.names[int(probs.top1)]
    conf = float(probs.top1conf)
    return label, conf


def draw_label_box(frame, x, y, w, h, label, conf):
    color = EMOTION_COLORS.get(label, (255, 255, 255))
    display_text = LABELS_ID.get(label, label.upper())
    text = f"{display_text} {conf*100:.0f}%"

    cv2.rectangle(frame, (x, y), (x + w, y + h), color, BOX_THICKNESS)

    (text_w, text_h), _ = cv2.getTextSize(text, FONT, FONT_SCALE, FONT_THICKNESS)
    label_y = max(0, y - text_h - 15)
    cv2.rectangle(frame, (x, label_y), (x + text_w + 12, label_y + text_h + 15), color, -1)

    brightness = sum(color) / 3
    text_color = (0, 0, 0) if brightness > 150 else (255, 255, 255)
    cv2.putText(frame, text, (x + 6, label_y + text_h + 5),
                FONT, FONT_SCALE, text_color, FONT_THICKNESS, cv2.LINE_AA)


def format_duration(seconds):
    seconds = int(seconds)
    h, rem = divmod(seconds, 3600)
    m, s = divmod(rem, 60)
    if h > 0:
        return f"{h:02d}:{m:02d}:{s:02d}"
    return f"{m:02d}:{s:02d}"


def draw_overlay(frame, elapsed_seconds, expression_totals):
    h_img, w_img = frame.shape[:2]

    detected = [(label, secs) for label, secs in expression_totals.items() if secs > 0]
    detected.sort(key=lambda item: item[1], reverse=True)

    lines = []
    for label, secs in detected:
        display_text = LABELS_ID.get(label, label.upper())
        lines.append((display_text, EMOTION_COLORS.get(label, (255, 255, 255)), format_duration(secs)))
    lines.append(("Time elapsed", (255, 255, 255), format_duration(elapsed_seconds)))

    box_w = 320
    box_h = OVERLAY_MARGIN * 2 + OVERLAY_LINE_HEIGHT * len(lines)
    box_x0 = OVERLAY_MARGIN
    box_y1 = h_img - OVERLAY_MARGIN
    box_y0 = box_y1 - box_h
    box_x1 = box_x0 + box_w

    overlay = frame.copy()
    cv2.rectangle(overlay, (box_x0, box_y0), (box_x1, box_y1), (0, 0, 0), -1)
    cv2.addWeighted(overlay, 0.55, frame, 0.45, 0, dst=frame)

    text_y = box_y0 + OVERLAY_MARGIN + 20
    for label_text, color, value_text in lines:
        line = f"{label_text}: {value_text}"
        cv2.putText(frame, line, (box_x0 + 14, text_y),
                    FONT, OVERLAY_FONT_SCALE, color, OVERLAY_FONT_THICKNESS, cv2.LINE_AA)
        text_y += OVERLAY_LINE_HEIGHT


def save_summary_report(path, start_time, expression_totals, timeline):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.chart import PieChart, BarChart, Reference

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Facial Expression Session Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Generated"
    ws["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A3"] = "Session start"
    ws["B3"] = start_time.strftime("%Y-%m-%d %H:%M:%S")
    ws["A4"] = "Total time elapsed (seconds)"
    ws["B4"] = round(timeline[-1]["elapsed"], 1) if timeline else 0

    header_row = 6
    ws.cell(row=header_row, column=1, value="Expression").font = Font(bold=True)
    ws.cell(row=header_row, column=2, value="Seconds").font = Font(bold=True)
    ws.cell(row=header_row, column=3, value="% of total face-time").font = Font(bold=True)
    for c in (1, 2, 3):
        ws.cell(row=header_row, column=c).fill = PatternFill("solid", start_color="DDDDDD")

    sorted_items = sorted(
        ((LABELS_ID.get(k, k.upper()), v) for k, v in expression_totals.items() if v > 0),
        key=lambda x: x[1], reverse=True
    )

    first_data_row = header_row + 1
    row = first_data_row
    for label, secs in sorted_items:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=round(secs, 1))
        row += 1
    last_data_row = row - 1

    if last_data_row >= first_data_row:
        total_cell = f"SUM(B{first_data_row}:B{last_data_row})"
        for r in range(first_data_row, last_data_row + 1):
            ws.cell(row=r, column=3, value=f"=B{r}/{total_cell}")
            ws.cell(row=r, column=3).number_format = "0.0%"

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 18

        pie = PieChart()
        pie.title = "Share of detected expression time"
        data = Reference(ws, min_col=2, min_row=header_row, max_row=last_data_row)
        cats = Reference(ws, min_col=1, min_row=first_data_row, max_row=last_data_row)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(cats)
        ws.add_chart(pie, "E6")

        bar = BarChart()
        bar.title = "Accumulated seconds per expression"
        bar.y_axis.title = "Seconds"
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        ws.add_chart(bar, "E22")
    else:
        ws.cell(row=first_data_row, column=1, value="No expressions detected during this session.")

    wb.save(path)
    print(f"Summary report saved: {path}")


def save_timeline_report(path, timeline):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws2 = wb.active
    ws2.title = "Timeline"
    expr_keys = list(LABELS_ID.keys())
    headers = ["Elapsed (s)"] + [LABELS_ID[k] for k in expr_keys] + ["Faces detected (this sample)"]
    ws2.append(headers)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", start_color="DDDDDD")

    for sample in timeline:
        row = [round(sample["elapsed"], 1)]
        for k in expr_keys:
            row.append(round(sample["deltas"].get(k, 0.0), 3))
        row.append(sample["face_count"])
        ws2.append(row)

    for i in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 14

    wb.save(path)
    print(f"Timeline report disimpan: {path}")


def save_reports(session_dir, start_time, expression_totals, timeline):
    os.makedirs(session_dir, exist_ok=True)
    save_summary_report(os.path.join(session_dir, "summary.xlsx"), start_time, expression_totals, timeline)
    save_timeline_report(os.path.join(session_dir, "timeline.xlsx"), timeline)


def list_cameras(max_index=10, backend=None):
    print("Scanning for available cameras...\n")
    found_any = False
    for idx in range(max_index):
        cap = cv2.VideoCapture(idx, backend) if backend is not None else cv2.VideoCapture(idx)
        if not cap.isOpened():
            cap.release()
            continue
        ret, frame = cap.read()
        if ret and frame is not None:
            h, w = frame.shape[:2]
            print(f"  [{idx}] OK - delivers frames at {w}x{h}")
            found_any = True
        else:
            print(f"  [{idx}] Opens but did not deliver a frame (likely not a real camera)")
        cap.release()
    if not found_any:
        print("\nNo working camera indices found. Try --backend dshow (Windows) or check Windows camera privacy settings.")
    print("\nUse the working index with: python live_webcam.py --camera <index>")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--camera", type=int, default=0, help="Camera index")
    parser.add_argument("--width", type=int, default=1920, help="Capture width")
    parser.add_argument("--height", type=int, default=1080, help="Capture height")
    parser.add_argument("--fullscreen", action="store_true", help="Start in true fullscreen")
    parser.add_argument(
        "--backend", choices=["auto", "dshow", "msmf"], default="auto",
        help="Windows capture backend. Try 'dshow' if an external webcam isn't picked up correctly with the default."
    )
    parser.add_argument(
        "--list-cameras", action="store_true",
        help="List all detected camera indices and their resolutions, then exit."
    )
    parser.add_argument(
        "--report-dir", default="report",
        help="Base folder for reports. Each run creates a session_<timestamp> subfolder here "
             "containing summary.xlsx and timeline.xlsx. Default: ./report"
    )
    args = parser.parse_args()

    backend_map = {
        "dshow": cv2.CAP_DSHOW,
        "msmf": cv2.CAP_MSMF,
    }
    backend = backend_map.get(args.backend)

    if args.list_cameras:
        list_cameras(backend=backend)
        sys.exit(0)

    cap = cv2.VideoCapture(args.camera, backend) if backend is not None else cv2.VideoCapture(args.camera)
    if not cap.isOpened():
        print(f"Could not open camera index {args.camera}.")
        print("Run 'python live_webcam.py --list-cameras' to see which indices actually work,")
        print("or try '--backend dshow' if you're on Windows and using an external/USB webcam.")
        return

    cap.set(cv2.CAP_PROP_FRAME_WIDTH, args.width)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, args.height)

    actual_w = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    actual_h = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    print(f"Requested {args.width}x{args.height}, camera is actually delivering {actual_w}x{actual_h}.")

    window_name = "Live Facial Expression Recognition"
    cv2.namedWindow(window_name, cv2.WINDOW_NORMAL)
    cv2.resizeWindow(window_name, actual_w, actual_h)

    fullscreen = args.fullscreen
    if fullscreen:
        cv2.setWindowProperty(window_name, cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_FULLSCREEN)

    print("Live detection dimulai. Tekan 'q' untuk keluar (saves report), 'f' untuk fullscreen, 'e' untuk membuat report tanpa keluar dari program ")

    start_dt = datetime.now()
    start_t = time.time()
    last_t = start_t

    expression_totals = defaultdict(float)
    timeline = []

    report_dir = os.path.join(args.report_dir, f"session_{start_dt.strftime('%Y%m%d_%H%M%S')}")

    consecutive_failures = 0
    max_consecutive_failures = 30

    try:
        while True:
            ret, frame = cap.read()
            if not ret or frame is None:
                consecutive_failures += 1
                if consecutive_failures >= max_consecutive_failures:
                    print("Camera stopped delivering frames. Exiting.")
                    break
                continue
            consecutive_failures = 0

            now = time.time()
            dt = now - last_t
            last_t = now
            elapsed = now - start_t

            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(
                gray, scaleFactor=1.1, minNeighbors=6, minSize=(80, 80)
            )

            h_img, w_img = frame.shape[:2]
            sample_deltas = defaultdict(float)

            for (x, y, w, h) in faces:
                x0, y0 = max(0, x - PADDING), max(0, y - PADDING)
                x1, y1 = min(w_img, x + w + PADDING), min(h_img, y + h + PADDING)
                face_crop = frame[y0:y1, x0:x1]
                label, conf = classify_face(face_crop)
                if label is None or label not in ALLOWED_EXPRESSIONS:
                    continue
                draw_label_box(frame, x0, y0, x1 - x0, y1 - y0, label, conf)

                if USE_PER_FACE_ACCUMULATION:
                    expression_totals[label] += dt
                    sample_deltas[label] += dt
                else:
                    sample_deltas[label] = dt

            if not USE_PER_FACE_ACCUMULATION:
                for label in sample_deltas:
                    expression_totals[label] += dt

            timeline.append({
                "elapsed": elapsed,
                "deltas": dict(sample_deltas),
                "face_count": len(faces),
            })

            draw_overlay(frame, elapsed, expression_totals)
            cv2.imshow(window_name, frame)

            key = cv2.waitKey(1) & 0xFF
            if key == ord("q"):
                break
            elif key == ord("f"):
                fullscreen = not fullscreen
                if fullscreen:
                    cv2.setWindowProperty(window_name, cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_FULLSCREEN)
                else:
                    cv2.setWindowProperty(window_name, cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_NORMAL)
                    cv2.resizeWindow(window_name, actual_w, actual_h)
            elif key == ord("e"):
                save_reports(report_dir, start_dt, expression_totals, timeline)
    finally:
        cap.release()
        cv2.destroyAllWindows()
        save_reports(report_dir, start_dt, expression_totals, timeline)


if __name__ == "__main__":
    main()
